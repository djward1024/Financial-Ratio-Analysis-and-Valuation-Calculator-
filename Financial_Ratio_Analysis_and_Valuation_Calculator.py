## This Python file is intended to create detailed financial ratio analysis, valuation metrics, and DCF calculations for a list of stocks via user-input. This will be exported into an excel file which displays the results in a table with the selected stocks in the located in the colums and the financial metrics and valuation metrics located in the rows. The goal of this script is to minimize human inputs to create seamless analysis. ##
## This script will use zero copying and pasting for learning purposes ##

## ----- Establish Libraries and Key Variables ----- ##

## Import Libraries ##

import os
import requests
import pandas as pd
import numpy as np
from datetime import datetime

# Ensures the script stays in its own folder
os.chdir(os.path.dirname(os.path.abspath(__file__)))

## API key ##

API_KEY = "1YKzdFA4grSB7HPSQCraS3nEVvFOWhmi"
BASE_URL = "https://financialmodelingprep.com/stable/"

## ----- Live Risk-Free Rate Fetcher ----- ##
## Pulls the current 10-Year Treasury yield from the U.S. Treasury API.
## Falls back to a hardcoded value if the request fails.

def get_risk_free_rate():
    """
    Fetches the most recent 10-Year Treasury yield from the U.S. Treasury API.
    Returns the rate as a decimal (e.g., 4.25% -> 0.0425).
    Falls back to 0.04046 if the live fetch fails.
    """
    fallback_rate = 0.04046
    try:
        # U.S. Treasury publishes daily yield curve data as XML
        url = "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml?data=daily_treasury_yield_curve&field_tdr_date_value=2025"
        response = requests.get(url, timeout=10)
        if response.status_code != 200:
            print(f"  Warning: Treasury API returned status {response.status_code}. Using fallback rate.")
            return fallback_rate

        # Parse the XML - the 10yr field is BC_10YEAR
        import xml.etree.ElementTree as ET
        root = ET.fromstring(response.content)
        ns = {'m': 'http://schemas.microsoft.com/ado/2007/08/dataservices/metadata',
              'd': 'http://schemas.microsoft.com/ado/2007/08/dataservices'}
        
        # Entries are in reverse chronological order; find the latest non-null 10yr value
        entries = root.findall('.//{http://www.w3.org/2005/Atom}entry')
        for entry in entries:
            bc_10yr = entry.find('.//d:BC_10YEAR', ns)
            if bc_10yr is not None and bc_10yr.text:
                rate = float(bc_10yr.text) / 100  # Convert from percent to decimal
                print(f"  Live 10-Year Treasury Yield: {rate:.4%}")
                return rate

        print("  Warning: Could not parse Treasury XML. Using fallback rate.")
        return fallback_rate

    except Exception as e:
        print(f"  Warning: Failed to fetch risk-free rate ({e}). Using fallback rate: {fallback_rate:.4%}")
        return fallback_rate

## Fetching Function ##

def get_financial_data(ticker, statement_type):

    #1 Define where the file should be (e.g., cache/AAPL_income-statement.json)

    if not os.path.exists("cache"):
        os.makedirs("cache")

    file_path = f"cache/{ticker}_{statement_type}.json"

    #2 Check: Do we already have this on the hard drive?

    if os.path.exists(file_path):
        print(f"  Loading {ticker} {statement_type} from local cache...")
        return pd.read_json(file_path)

    #3 If not in cache, go to internet

    print(f"  Fetching {ticker} {statement_type} from FMP API...")
    url = f"{BASE_URL}{statement_type}?symbol={ticker}&apikey={API_KEY}"
    if "statement" in statement_type:
        url += "&limit=5"

    response = requests.get(url)

    # Check if the server actually received a "Success" code

    if response.status_code != 200:
        print(f"  !!! Error: Server returned status {response.status_code} for {ticker}")
        return None

    # Check if the response body is empty or not JSON

    if not response.text.strip():
        print(f"  !!! Error: Received empty response for {ticker}")
        return None

    try:
        json_data = response.json()
    except requests.exceptions.JSONDecodeError:
        print(f"  !!! Error: Could not parse JSON for {ticker}. The API likely sent an HTML error page.")
        return None

    # Legacy error message fix

    if isinstance(json_data, dict):
        if "Error Message" in json_data or "error" in json_data:
            print(f"  !!! API ERROR for {ticker}: {json_data.get('Error Message', 'Unknown Error')}")
            print(f"  Debug URL: {url}")
            return None
        df = pd.DataFrame([json_data])
    else:
        if not json_data:
            print(f"  Warning: No data returned for {ticker}")
            return None
        df = pd.DataFrame(json_data)

    #4 Save the data so we don't have to pay or wait for it again

    df.to_json(file_path)

    return df


## ----- Calculation Function ----- ##

def calculate_metrics(inc, bal, cf, quote, risk_free_rate):

    # Store results

    results = {}

    #1 Valuation Metrics

    mkt_cap = quote['marketCap'].iloc[0]
    price = quote['price'].iloc[0]
    enterprise_value = mkt_cap + bal['totalDebt'].iloc[0] - bal['cashAndCashEquivalents'].iloc[0]
    results['P/E Ratio'] = mkt_cap / inc['netIncome'].iloc[0]
    results['P/B Ratio'] = mkt_cap / bal['totalStockholdersEquity'].iloc[0]
    ebitda = inc['ebitda'].iloc[0] if 'ebitda' in inc.columns else inc['operatingIncome'].iloc[0]
    results['EV/EBITDA'] = enterprise_value / ebitda if ebitda != 0 else 0

    #2 Profitability

    results['Gross Margin'] = inc['grossProfit'].iloc[0] / inc['revenue'].iloc[0]
    results['Operating Margin'] = inc['operatingIncome'].iloc[0] / inc['revenue'].iloc[0]
    results['Net Margin'] = inc['netIncome'].iloc[0] / inc['revenue'].iloc[0]

    #3 Return Ratios

    results['ROE'] = inc['netIncome'].iloc[0] / bal['totalStockholdersEquity'].iloc[0]
    results['ROA'] = inc['netIncome'].iloc[0] / bal['totalAssets'].iloc[0]
    invested_capital = bal['totalDebt'].iloc[0] + bal['totalStockholdersEquity'].iloc[0]
    results['ROIC'] = inc['operatingIncome'].iloc[0] / invested_capital

    #4 Liquidity & Solvency Ratios

    raw_interest = inc['interestExpense'].iloc[0] if 'interestExpense' in inc.columns else 0
    int_exp = abs(raw_interest)
    results['Current Ratio'] = bal['totalCurrentAssets'].iloc[0] / bal['totalCurrentLiabilities'].iloc[0]
    results['Debt-to-Equity'] = bal['totalDebt'].iloc[0] / bal['totalStockholdersEquity'].iloc[0]
    # Traditionally Interest Coverage = EBIT / Interest Expense
    results['Interest Coverage'] = inc['operatingIncome'].iloc[0] / int_exp if int_exp > 0 else "N/A (Cash Rich)"

    #5 Turnover Ratios

    inv = bal['inventory'].iloc[0]
    results['Days Sales Outstanding'] = (bal['netReceivables'].iloc[0] / inc['revenue'].iloc[0]) * 365
    results['Inventory Turnover'] = inc['costOfRevenue'].iloc[0] / inv if inv > 0 else 0

    #6 Dividend Sustainability

    # Extract Dividends Paid

    div_paid = abs(cf['dividendsPaid'].iloc[0]) if 'dividendsPaid' in cf.columns else 0
    net_income = inc['netIncome'].iloc[0]

    # Payout Ratio (Earnings Basis)
    results['Dividend Payout Ratio'] = div_paid / net_income if net_income > 0 else 0

    # Cash Flow Coverage
    fcf_current = cf['operatingCashFlow'].iloc[0] + cf['capitalExpenditure'].iloc[0]
    results['Dividend Coverage (FCF)'] = fcf_current / div_paid if div_paid > 0 else "No Dividend"

    # Dividend Yield
    if 'sharesOutstanding' in quote.columns:
        shares_out = quote['sharesOutstanding'].iloc[0]
    elif 'weightedAverageShsOut' in inc.columns:
        shares_out = inc['weightedAverageShsOut'].iloc[0]
    else:
        shares_out = 1  # Fallback to avoid division by 0
    div_per_share = div_paid / shares_out if shares_out > 0 else 0
    results['Dividend Yield'] = div_per_share / price if price > 0 else 0

    #7 DCF Analysis
    ## KEY CHANGES:
    ## 1. Risk-free rate is now passed in as a live variable (fetched from U.S. Treasury)
    ## 2. Removed the decay_factor — it was compounding too aggressively with the growth cap
    ##    and producing unrealistically low terminal values.
    ## 3. Revenue CAGR is used for FCF growth estimate but is capped and then staged:
    ##    Years 1-3 use full (capped) growth, Years 4-5 step down toward terminal growth.
    ##    This reflects a more realistic "fade" without the harsh decay_factor math.
    ## 4. FCF floor: if normalized FCF is negative (pre-profit companies like RIVN),
    ##    the DCF returns $0 rather than a nonsensical negative intrinsic value.

    # Establish Variables

    # CAGR Revenue Growth Rate

    rev_now = inc['revenue'].iloc[0]
    rev_past = inc['revenue'].iloc[-1]
    years = len(inc) - 1
    growth_rate = (rev_now / rev_past) ** (1 / years) - 1 if years > 0 else 0.05
    # Cap growth at 25% to prevent hyper-inflation of value (raised slightly from 20%)
    growth_rate = min(growth_rate, 0.25)

    # Estimate Cost of Debt (After-Tax)

    tax_rate = inc['incomeTaxExpense'].iloc[0] / inc['incomeBeforeTax'].iloc[0] \
        if inc['incomeBeforeTax'].iloc[0] != 0 else 0.21
    # Clamp tax rate to a realistic range (0% - 40%)
    tax_rate = max(0.0, min(0.40, tax_rate))

    interest_expense = abs(inc['interestExpense'].iloc[0]) if 'interestExpense' in inc.columns else 0
    total_debt = bal['totalDebt'].iloc[0]
    pre_tax_cost_of_debt = interest_expense / total_debt if total_debt > 0 else 0.05
    cost_of_debt = pre_tax_cost_of_debt * (1 - tax_rate)

    # Estimate Cost of Equity (CAPM)
    # Risk-free rate is now LIVE from U.S. Treasury (passed in as argument)
    # ERP: 5.5% is a reasonable long-run estimate (Damodaran consensus)

    Equity_Risk_Premium = 0.0550
    beta = quote['beta'].iloc[0] if 'beta' in quote.columns and pd.notnull(quote['beta'].iloc[0]) else 1.0
    cost_of_equity = risk_free_rate + (beta * Equity_Risk_Premium)

    # WACC

    weight_of_equity = mkt_cap / (mkt_cap + total_debt) if (mkt_cap + total_debt) > 0 else 1.0
    weight_of_debt = total_debt / (mkt_cap + total_debt) if (mkt_cap + total_debt) > 0 else 0.0
    wacc = (cost_of_equity * weight_of_equity) + (cost_of_debt * weight_of_debt)

    # Terminal Growth: lower of growth_rate or 3% (slightly more generous than 2.5%)
    # 3% roughly matches long-run nominal GDP growth

    terminal_growth = min(growth_rate, 0.03)

    # Ensure WACC > terminal growth to prevent math blowup

    if wacc <= terminal_growth:
        wacc = terminal_growth + 0.01

    # Normalized FCF Base (3-year average)

    fcf_series = cf['operatingCashFlow'].head(3) + cf['capitalExpenditure'].head(3)
    fcf = fcf_series.mean()

    # FCF Floor: if the company is burning cash, DCF is not meaningful

    if fcf <= 0:
        results['DCF Intrinsic Value'] = 0
        results['Margin of Safety'] = 0
    else:
        # Staged Growth Projection (Years 1-5)
        # Years 1-3: Full capped growth rate
        # Years 4-5: Linearly interpolate between growth_rate and terminal_growth
        # This replaces the harsh decay_factor with a more intuitive "fade"

        dcf_val = 0
        future_fcf = fcf
        for i in range(1, 6):
            if i <= 3:
                yearly_growth = growth_rate
            else:
                # Fade from growth_rate toward terminal_growth over years 4-5
                fade_steps = i - 3          # 1 or 2
                total_fade = 2              # 2 steps to fade
                yearly_growth = growth_rate - (fade_steps / total_fade) * (growth_rate - terminal_growth)
            future_fcf *= (1 + yearly_growth)
            dcf_val += future_fcf / (1 + wacc) ** i

        # Terminal Value

        terminal_val = (future_fcf * (1 + terminal_growth)) / (wacc - terminal_growth)
        dcf_val += terminal_val / (1 + wacc) ** 5

        # Adjust for Net Debt to get Equity Value per Share

        net_debt = total_debt - bal['cashAndCashEquivalents'].iloc[0]
        share_col = [col for col in quote.columns if 'shares' in col.lower()]

        if share_col:
            shares = quote[share_col[0]].iloc[0]
        elif 'weightedAverageShsOut' in inc.columns:
            shares = inc['weightedAverageShsOut'].iloc[0]
        else:
            shares = 0

        intrinsic_value = (dcf_val - net_debt) / shares if shares > 0 else 0

        if intrinsic_value > 0:
            results['Margin of Safety'] = (intrinsic_value - price) / intrinsic_value
        else:
            results['Margin of Safety'] = 0

        results['DCF Intrinsic Value'] = max(intrinsic_value, 0)

    #8 Bankruptcy & Manipulation Scores

    # Z-Score Inputs

    total_assets = bal['totalAssets'].iloc[0]
    working_cap = bal['totalCurrentAssets'].iloc[0] - bal['totalCurrentLiabilities'].iloc[0]
    retained_earnings = bal['retainedEarnings'].iloc[0] if 'retainedEarnings' in bal.columns else 0
    ebit = inc['operatingIncome'].iloc[0]
    total_liab = bal['totalLiabilities'].iloc[0]
    sales = inc['revenue'].iloc[0]

    A = working_cap / total_assets
    B = retained_earnings / total_assets
    C = ebit / total_assets
    D = mkt_cap / total_liab if total_liab > 0 else 0
    E = sales / total_assets

    # Z-Score Calculation

    results['Altman Z-Score'] = (1.2 * A) + (1.4 * B) + (3.3 * C) + (0.6 * D) + (1.0 * E)

    # Beneish M-Score Inputs (At Least 2 Years of Data)

    if len(inc) > 1 and len(bal) > 1:
        s_t, s_prev = inc['revenue'].iloc[0], inc['revenue'].iloc[1]
        rec_t, rec_prev = bal['netReceivables'].iloc[0], bal['netReceivables'].iloc[1]
        cogs_t, cogs_prev = inc['costOfRevenue'].iloc[0], inc['costOfRevenue'].iloc[1]
        ca_t, ca_prev = bal['totalCurrentAssets'].iloc[0], bal['totalCurrentAssets'].iloc[1]
        ta_t, ta_prev = bal['totalAssets'].iloc[0], bal['totalAssets'].iloc[1]
        dep_t, dep_prev = inc['depreciationAndAmortization'].iloc[0], inc['depreciationAndAmortization'].iloc[1]
        ppe_t, ppe_prev = bal['propertyPlantEquipmentNet'].iloc[0], bal['propertyPlantEquipmentNet'].iloc[1]
        sga_t, sga_prev = inc['sellingGeneralAndAdministrativeExpenses'].iloc[0], inc['sellingGeneralAndAdministrativeExpenses'].iloc[1]
        lt_debt_t, lt_debt_prev = bal['longTermDebt'].iloc[0], bal['longTermDebt'].iloc[1]
        ni_t = inc['netIncome'].iloc[0]
        cfo_t = cf['operatingCashFlow'].iloc[0]

        # Indices

        dsri = (rec_t / s_t) / (rec_prev / s_prev) if s_prev > 0 and rec_prev > 0 else 1
        gmi = ((s_prev - cogs_prev) / s_prev) / ((s_t - cogs_t) / s_t) if s_t > 0 else 1
        aqi = (1 - (ca_t + ppe_t) / ta_t) / (1 - (ca_prev + ppe_prev) / ta_prev) if ta_prev > 0 else 1
        sgi = s_t / s_prev if s_prev > 0 else 1
        depi = (dep_prev / (ppe_prev + dep_prev)) / (dep_t / (ppe_t + dep_t)) if (ppe_t + dep_t) > 0 else 1
        sgai = (sga_t / s_t) / (sga_prev / s_prev) if s_prev > 0 else 1
        lvgi = ((lt_debt_t + bal['totalCurrentLiabilities'].iloc[0]) / ta_t) / \
               ((lt_debt_prev + bal['totalCurrentLiabilities'].iloc[1]) / ta_prev) if ta_prev > 0 else 1
        tata = (ni_t - cfo_t) / ta_t

        # M-Score Calculation

        results['Beneish M-Score'] = (-4.84 + (0.92 * dsri) + (0.528 * gmi) + (0.404 * aqi) +
                                      (0.892 * sgi) + (0.115 * depi) - (0.172 * sgai) +
                                      (4.679 * tata) - (0.327 * lvgi))
    else:
        results['Beneish M-Score'] = "N/A (Need 2yr Data)"

    #9 Portfolio Ranking Logic

    score = 0

    # Valuation Score (Max 35 points)
    mos = results.get('Margin of Safety', 0)
    score += max(0, min(35, (mos / 0.20) * 35))

    # Solvency Score (Max 25 points)
    z_score = results.get('Altman Z-Score', 0)
    score += max(0, min(25, (z_score / 6.0) * 25))

    # Profitability Score (Max 20 points)
    roic = results.get('ROIC', 0)
    score += max(0, min(20, (roic / 0.20) * 20))

    # Dividend/Cash Safety (Max 20 points)
    div_cov = results.get('Dividend Coverage (FCF)', 0)
    if isinstance(div_cov, (int, float)):
        score += max(0, min(20, (div_cov / 2.0) * 20))
    else:
        score += 10  # Neutral points for non-dividend payers

    results['FINAL RANK SCORE'] = score

    return results


## ----- Execution Phase ----- ##

# Fetch live risk-free rate ONCE before the ticker loop

print("\nFetching live 10-Year Treasury yield...")
RISK_FREE_RATE = get_risk_free_rate()

# Ask user their list of stocks

ticker_input = input("\nPlease enter the stock tickers (e.g., AAPL, DD, UNH): ").upper()
ticker_list = [t.strip() for t in ticker_input.split(",")]

# "Master Directory"

all_company_results = {}

# Start the loop and process each input

for ticker in ticker_list:
    print(f"\nRetrieving data for: {ticker}...")
    income = get_financial_data(ticker, "income-statement")
    balance = get_financial_data(ticker, "balance-sheet-statement")
    cash_flow = get_financial_data(ticker, "cash-flow-statement")
    quote = get_financial_data(ticker, "quote")

    if income is not None and balance is not None and cash_flow is not None and quote is not None:
        all_company_results[ticker] = calculate_metrics(income, balance, cash_flow, quote, RISK_FREE_RATE)
        print(f"  Analysis complete for {ticker}")
    else:
        print(f"  Skipping {ticker} due to missing data.")

# Create timestamp

timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
filename = f"Stock_Analysis_{timestamp}.xlsx"

## ----- Excel Export & Formatting ----- ##

if all_company_results:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

    final_df = pd.DataFrame(all_company_results)
    final_df.to_excel(filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Style definitions

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")

    # Format Headers (Tickers) and Index (Metric Names)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for cell in ws['A']:
        cell.font = Font(bold=True)

    # Format Data Rows

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2):
        for cell in row:
            metric_name = ws.cell(row=cell.row, column=1).value
            if metric_name is None:
                continue
            if 'FINAL RANK SCORE' in metric_name:
                cell.number_format = '0'
            elif any(pct in metric_name for pct in ['Margin', 'RO', 'Safety', 'Dividend Yield', 'Dividend Payout']):
                cell.number_format = '0.00%'
            elif 'Value' in metric_name:
                cell.number_format = '$#,##0.00'
            else:
                cell.number_format = '0.00'

    ## ----- Conditional Formatting ----- ##
    ## REDESIGNED: Each metric is now explicitly classified as either
    ## "higher is better" (standard green-to-red scale) or
    ## "lower is better" (reversed red-to-green scale), or gets a custom rule.
    ##
    ## Color scales:
    ##   Standard  (high=green): min=red  F8696B, mid=yellow FFEB84, max=green 63BE7B
    ##   Reversed  (low=green):  min=green 63BE7B, mid=yellow FFEB84, max=red F8696B
    ##
    ## Individual threshold rules use:
    ##   green_fill = C6EFCE (light green)
    ##   red_fill   = FFC7CE (light red)

    # Re-usable fill objects
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")

    # Standard color scale: low=red, high=green (higher is better)
    standard_scale = ColorScaleRule(
        start_type='min',        start_color="F8696B",
        mid_type='percentile',   mid_value=50, mid_color="FFEB84",
        end_type='max',          end_color="63BE7B"
    )

    # Reversed color scale: low=green, high=red (lower is better)
    reversed_scale = ColorScaleRule(
        start_type='min',        start_color="63BE7B",
        mid_type='percentile',   mid_value=50, mid_color="FFEB84",
        end_type='max',          end_color="F8696B"
    )

    for row_idx in range(2, ws.max_row + 1):
        metric_label = str(ws.cell(row=row_idx, column=1).value).strip()
        col_start = get_column_letter(2)
        col_end   = get_column_letter(ws.max_column)
        cell_range = f"{col_start}{row_idx}:{col_end}{row_idx}"

        # ── FINAL RANK SCORE ── Bold + standard scale (higher is better)
        if 'FINAL RANK SCORE' in metric_label:
            ws.conditional_formatting.add(cell_range, standard_scale)
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True, size=12)

        # ── MARGIN OF SAFETY ── Custom thresholds
        # > 20% = green (good upside), < 0% = red (overvalued)
        elif 'Margin of Safety' in metric_label:
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=['0.2'], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan',    formula=['0'],   fill=red_fill))

        # ── ALTMAN Z-SCORE ── Custom thresholds
        # > 3.0 = safe (green), 1.8–3.0 = grey zone (yellow), < 1.8 = distress (red)
        elif 'Altman Z-Score' in metric_label:
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=['3.0'], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan',    formula=['1.8'], fill=red_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='between',     formula=['1.8', '3.0'], fill=yellow_fill))

        # ── BENEISH M-SCORE ── Custom thresholds
        # < -2.22 = likely not manipulating (green), > -2.22 = manipulation risk (red)
        elif 'Beneish M-Score' in metric_label:
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan',    formula=['-2.22'], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=['-2.22'], fill=red_fill))

        # ── DIVIDEND PAYOUT RATIO ── Custom thresholds
        # < 50% = sustainable (green), > 80% = danger zone (red)
        elif 'Dividend Payout Ratio' in metric_label:
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan',    formula=['0.5'], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=['0.8'], fill=red_fill))

        # ── CURRENT RATIO ── Custom thresholds
        # 1.5–3.0 = healthy (green), < 1.0 = liquidity risk (red), > 5.0 = excess idle cash (yellow)
        elif 'Current Ratio' in metric_label:
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='between', formula=['1.5', '3.0'], fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan', formula=['1.0'], fill=red_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=['5.0'], fill=yellow_fill))

        # ── INTEREST COVERAGE ── Custom thresholds
        # > 5x = very safe (green), < 1.5x = danger (red)
        elif 'Interest Coverage' in metric_label:
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='greaterThan', formula=['5'],   fill=green_fill))
            ws.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan',    formula=['1.5'], fill=red_fill))

        # ── METRICS WHERE LOWER IS BETTER (reversed color scale) ──
        # P/E, P/B, EV/EBITDA: cheaper valuation = greener
        # Debt-to-Equity: less debt = greener
        # Days Sales Outstanding: faster collection = greener
        elif any(m in metric_label for m in ['P/E', 'P/B', 'EV/EBITDA', 'Debt-to-Equity', 'Days Sales Outstanding']):
            ws.conditional_formatting.add(cell_range, reversed_scale)

        # ── METRICS WHERE HIGHER IS BETTER (standard color scale) ──
        # Margins, Returns, Turnover, Dividend Coverage, Dividend Yield, Inventory Turnover
        elif any(m in metric_label for m in [
            'Gross Margin', 'Operating Margin', 'Net Margin',
            'ROE', 'ROA', 'ROIC',
            'Inventory Turnover', 'Dividend Coverage', 'Dividend Yield'
        ]):
            ws.conditional_formatting.add(cell_range, standard_scale)

    # Final column widths

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    wb.save(filename)
    print(f"\nSuccess! Report saved as: {filename}")

else:
    print("\nNo data was retrieved. Please check your tickers and API key.")